# extractor.py
# Dynamically parse the downloaded Table 6 Excel file and extract
# Real Estate and Household Equipment data from each year tab.

import logging
import openpyxl

import config

logger = logging.getLogger(__name__)


class TNFADataParser:
    """
    Parses the downloaded Table 6 "Assets Structure for Households Sector"
    spreadsheet from the Taiwan DGBAS site.

    The parser is fully dynamic:
      - Sheet tabs are year numbers (e.g. 2023, 2022, 2021...).
      - It scans each tab to find the Amount and Composition columns
        under the FIRST asset group header.
      - It searches for item rows by partial text matching in column B.
      - No row/column positions are hardcoded.
    """

    def __init__(self):
        self.logger = logger

    # ─────────────────────────────────────────────────────────────────────
    # Internal helpers
    # ─────────────────────────────────────────────────────────────────────

    def _get_year_sheets(self, wb):
        """
        Return sheet names that are valid years, sorted descending (latest first).
        """
        year_sheets = []
        for name in wb.sheetnames:
            try:
                year = int(name.strip())
                if 1900 <= year <= 2100:
                    year_sheets.append((year, name))
            except (ValueError, TypeError):
                continue

        year_sheets.sort(key=lambda x: x[0], reverse=True)
        self.logger.info(
            f'Year tabs found: {[y for y, _ in year_sheets]}'
        )
        return year_sheets

    def _find_amount_composition_cols(self, ws):
        """
        Scan the header rows to find the Amount and Composition column
        indices for the FIRST asset group.

        The structure is:
          Row 2: merged headers for asset groups spanning 2 cols each
          Row 3: "Amount (100 Million NT$)" | "Composition (%)" repeated

        Returns (amount_col, composition_col) — 1-based column indices.
        """
        # Scan rows 2-5 for "Amount" and "Composition" sub-headers
        for row_idx in range(1, min(ws.max_row + 1, 10)):
            amount_col = None
            composition_col = None

            for col_idx in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                val_str = str(val).strip().lower()

                if 'amount' in val_str and amount_col is None:
                    amount_col = col_idx
                elif 'composition' in val_str and amount_col is not None and composition_col is None:
                    composition_col = col_idx
                    break

            if amount_col and composition_col:
                self.logger.info(
                    f'Found headers at row {row_idx}: '
                    f'Amount=col {amount_col}, Composition=col {composition_col}'
                )
                return amount_col, composition_col

        # Fallback: columns C=3 (Amount) and D=4 (Composition)
        self.logger.warning(
            'Could not find Amount/Composition headers dynamically — '
            'falling back to columns C(3) and D(4)'
        )
        return 3, 4

    def _find_item_row(self, ws, item_text, start_row=1, end_row=None):
        """
        Search column B (col 2) for a row whose text contains item_text
        (case-insensitive partial match).

        Returns the row index or None.
        """
        end_row = end_row or ws.max_row
        item_lower = item_text.strip().lower()

        for row_idx in range(start_row, end_row + 1):
            cell_val = ws.cell(row=row_idx, column=2).value
            if cell_val is None:
                continue
            cell_str = str(cell_val).strip().lower()
            if item_lower in cell_str:
                return row_idx

        return None

    def _extract_sheet_data(self, ws, sheet_year):
        """
        Extract Real Estate and Household Equipment data from one year tab.

        Returns:
            {
                'REALESTATE': {'AMOUNT': value, 'COMPOSITION': value},
                'HOUSEHOLDEQUIPMENT': {'AMOUNT': value, 'COMPOSITION': value},
            }
            or None if extraction fails.
        """
        amount_col, composition_col = self._find_amount_composition_cols(ws)

        result = {}
        for key, search_text in config.EXTRACT_ITEMS.items():
            row_idx = self._find_item_row(ws, search_text, start_row=4)
            if row_idx is None:
                self.logger.warning(
                    f'  [{sheet_year}] Item not found: "{search_text}"'
                )
                continue

            amount_val = ws.cell(row=row_idx, column=amount_col).value
            comp_val = ws.cell(row=row_idx, column=composition_col).value

            # Validate numeric values
            if amount_val is not None and not isinstance(amount_val, str):
                result[key] = {
                    'AMOUNT': amount_val,
                    'COMPOSITION': comp_val if comp_val is not None and not isinstance(comp_val, str) else None,
                }
                self.logger.debug(
                    f'  [{sheet_year}] {key}: '
                    f'Amount={amount_val}, Composition={comp_val}'
                )
            else:
                self.logger.warning(
                    f'  [{sheet_year}] Non-numeric value for "{search_text}": {amount_val}'
                )

        return result if result else None

    # ─────────────────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────────────────

    def parse_excel(self, file_path):
        """
        Parse the downloaded Table 6 .xlsx file and return:

            {
                'data': {
                    year: {
                        'REALESTATE': {'AMOUNT': val, 'COMPOSITION': val},
                        'HOUSEHOLDEQUIPMENT': {'AMOUNT': val, 'COMPOSITION': val},
                    },
                    ...
                },
                'years': [2019, 2020, 2021, 2022, 2023],
                'min_year': 2019,
                'max_year': 2023,
            }
        """
        self.logger.info(f'Parsing Table 6 Excel: {file_path}')

        wb = openpyxl.load_workbook(file_path, data_only=True)
        year_sheets = self._get_year_sheets(wb)

        if not year_sheets:
            wb.close()
            raise RuntimeError('No valid year tabs found in the workbook')

        # Determine which sheets to process based on TARGET_YEAR
        target = config.TARGET_YEAR
        if target is None:
            # Default: latest year only
            sheets_to_process = [year_sheets[0]]
            self.logger.info(f'TARGET_YEAR=None -> extracting latest year: {sheets_to_process[0][0]}')
        elif target == 'ALL':
            sheets_to_process = year_sheets
            self.logger.info(f'TARGET_YEAR=ALL -> extracting all {len(sheets_to_process)} years')
        else:
            # Specific year
            sheets_to_process = [(y, n) for y, n in year_sheets if y == int(target)]
            if not sheets_to_process:
                wb.close()
                raise RuntimeError(f'Year tab {target} not found in workbook')
            self.logger.info(f'TARGET_YEAR={target} -> extracting specific year')

        all_data = {}
        for year, sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            self.logger.info(f'Processing sheet: {sheet_name}')

            sheet_data = self._extract_sheet_data(ws, year)
            if sheet_data:
                all_data[year] = sheet_data

        wb.close()

        if not all_data:
            raise RuntimeError('No data extracted from any year tab')

        sorted_years = sorted(all_data.keys())

        self.logger.info(f'Extraction complete:')
        self.logger.info(f'  Years: {sorted_years}')
        self.logger.info(f'  Items per year: {list(config.EXTRACT_ITEMS.keys())}')

        return {
            'data': all_data,
            'years': sorted_years,
            'min_year': sorted_years[0],
            'max_year': sorted_years[-1],
        }
